
import pandas as pd
import openpyxl
import os


def board_number(name):
    
    # Trækker tallet ud af mappenavnet f.eks. board_10 → 10
    
    return int(name.split("_")[1])

def load_ground_truth(excel_path):
    """
    Læser ground truth Excel_filen og konvertere den
    til et dictionary --->  {(borad_name, tile_file): label}.
    """
    
    # workbook er selve Excel-filen, der indeholder alle ark
    
    wb = openpyxl.load_workbook(excel_path)
    
    ws = wb.active     # worksheet --> det aktiver ark inde i Excel-filen
    
    # Læs klonnenavn for første række
    
    headers = [cell.value for cell in ws[1]]
    
    lookup= {}
    
    # Gå gennem alle rækker efter første række
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        board_name = row[0]
        if not board_name:
            continue
        
    # Gå gennem alle tile-klonner og byg dictionary "lookup"
    
        for col_idx, label in enumerate(row[1:], start=1):
            tile_col = headers[col_idx]       # F.eks "tile_0_0"
            
            if tile_col is None:
                continue
            
            tile_file = tile_col + ".jpg"     # F.eks "tile_0_0.jpg"
            
            if label :
                clean_label = str(label).strip().capitalize()
                lookup[(board_name, tile_file)] = clean_label
            
    return lookup

def merge_features_with_labels(features_csv, excel_path, output_csv):
    """ 
    Kobler features.csv med ground truth Excel filen og 
    erstatter alle unknown labels med de rigtige terræntyper.
    """
    
    # Indlæs features CSV
    df = pd.read_csv(features_csv)
    
    # Indlæs ground truth labels fra Excel-filen
    ground_truth = load_ground_truth(excel_path)
    
    # Estatter Unknown med rigtige label, via opslag i ground_truth.
    
    def find_label(row):
        key = (row["board"], row["tile_file"])
        return ground_truth.get(key, "unknown")

    df["label"] = df.apply(find_label, axis=1)
    
    # Fjern board-klonnen- bruges kun til opslag
    df = df.drop(columns=["board"])
    
    # Gem den opdaterede csv med rigtige labels
    df.to_csv(output_csv, index=False)
    
    print(f"Saved: {output_csv}")
    print(f"Total_tiles: {len(df)}")
    print(f"\nLabel distribution:")
    print(df["label"].value_counts())
    
    
    
    


    
    
    

            
        
        
        
        
        